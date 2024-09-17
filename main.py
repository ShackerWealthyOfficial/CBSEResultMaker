from openpyxl import Workbook as wb

roll_length = 8
gen_length = 11
name_length = [13, 65]
marks_length = 64
marks_spacing = 12
roll = []
gen = []
name = []
subs = []
grades = []
state = []
marks = []
comp_sub = []
total_subs = []

new_list = []

def remove_header(file):
    st = ""
    j = 0
    for i in file.readlines():
        if j != 0 or i[0] == '-':
            if j < 5:
                j += 1
                continue
            else:
                j = 0
        elif '*****' in i or i.isspace():
            pass
        elif i[0].isdigit() or i[0].isspace():
            st += i
    file2 = open('file2.txt', 'w')
    file2.write(st)
    file2.close()

def divide(file):
    a = file.readlines()

    for i in a:
        if i[0].isdigit() == True:
            l = 0
            subs.append([])
            grades.append([])
            marks.append([])

            roll.append(i[:roll_length:])
            gen.append(i[gen_length])
            name.append(i[name_length[0]:name_length[1]:].rstrip())
            for j in range(name_length[1], len(i)):
                if i[j].isdigit() and i[j-1].isspace():
                    subs[-1].append(i[j:j+3:])
                    l = j + 3
                elif i[j] in 'ABCDEF' and i[j+1].isdigit(): 
                    grades[-1].append(i[j:j+2:])
                elif i[j:j+4:] == 'COMP' or i[j:j+4:] == 'PASS':
                    state.append(i[j:j+4:])

            if i.strip()[-1].isdigit():
                comp_sub.append(i.strip()[-3::])
            else:
                comp_sub.append(None)

        else:
            temp = i.split()
            for k in range(0, len(temp), 2):
                marks[-1].append([temp[k], temp[k+1]])

def total_subjects():
    global total_subs
    l = []
    for i in subs:
        for j in i:
            if j not in l:
                l.append(j)

    total_subs = l

def arrange():
    new_list.append(roll)
    new_list.append(gen)
    new_list.append(name)
    new_list.append(subs)
    new_list.append(grades)
    new_list.append(state)
    new_list.append(marks)

def xlsx_create():
    wbook = wb()
    sheet = wbook.active

    temp = False
    temp2 = False

    cell_roll = sheet.cell(row = 1, column = 1)
    cell_roll.value = 'Roll No.'
    cell_roll = sheet.cell(row = 1, column = 2)
    cell_roll.value = 'Gender'
    cell_roll = sheet.cell(row = 1, column = 3)
    cell_roll.value = 'Name'
    sheet.column_dimensions['C'].width = name_length[1] - name_length[0]

    for r in range(0, len(roll)):
        cell_r = sheet.cell(row = r + 2, column = 1)
        cell_r.value = roll[r]
        cell_gen = sheet.cell(row = r + 2, column = 2)
        cell_gen.value = gen[r]
        cell_name = sheet.cell(row = r + 2, column = 3)
        cell_name.value = name[r]

    for i in range(0, len(total_subs) * 2, 2):
        sheet.column_dimensions[chr(ord('D') + (i // 2))].width = marks_spacing
        temp = 5 + i
        temp2 = (ord('D') + (i // 2))
        cell_sub1 = sheet.cell(row = 1, column = 4 + i)
        cell_sub1.value = (str(total_subs[i//2]) + '_marks')
        cell_sub2 = sheet.cell(row = 1, column = 5 + i)
        cell_sub2.value = str(total_subs[i//2]) + '_grades'
        
        for j in range(0, len(subs)):
            for k in range(0, len(subs[0])):
                if (cell_sub1.value[:3:]) == subs[j][k]:
                    cell_marks = sheet.cell(row = 2 + j, column = 4 + i)
                    cell_marks.value = marks[j][k][0]
                    cell_grades = sheet.cell(row = 2 + j, column = 5 + i)
                    cell_grades.value = marks[j][k][1]

    cell_grade1 = sheet.cell(row = 1, column = temp + 1)
    cell_grade1.value = 'Grade 1'
    cell_grade2 = sheet.cell(row = 1, column = temp + 2)
    cell_grade2.value = 'Grade 2'
    cell_grade3 = sheet.cell(row = 1, column = temp + 3)
    cell_grade3.value = 'Grade 3'
    
    cell_result = sheet.cell(row = 1, column = temp + 4)
    cell_result.value = 'Result'
    cell_comp_sub = sheet.cell(row = 1, column = temp + 5)
    cell_comp_sub.value = 'COMP SUB'

    for i in range(0, len(state)):
        cell_grades1 = sheet.cell(row = 2 + i, column = temp + 1)
        cell_grades1.value = grades[i][0]
        cell_grades2 = sheet.cell(row = 2 + i, column = temp + 2)
        cell_grades2.value = grades[i][1]
        cell_grades3 = sheet.cell(row = 2 + i, column = temp + 3)
        cell_grades3.value = grades[i][2]
        cell_res = sheet.cell(row = 2 + i, column = temp + 4)
        cell_res.value = state[i]
        cell_com = sheet.cell(row = 2 + i, column = temp + 5)
        if comp_sub[i]:
            cell_com.value = comp_sub[i]

    wbook.save('C:\\Users\\kv\\Desktop\\AD\\Project\\test.xlsx')

file = open('sample.txt', 'r')
remove_header(file)
file2 = open('file2.txt', 'r')
divide(file2)

arrange()
total_subjects()
xlsx_create()

file = open('sample.txt', 'r')
remove_header(file)
file2 = open('file2.txt', 'r')
divide(file2)

file.close()
file2.close()

file.close()
file2.close()


